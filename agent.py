"""Self-editing skill-memory agent on Gemini.

Multi-user. Per-user state lives in Postgres (see store.py + schema.sql).

Flow per turn:
  1. Read every skill (name+description+body+tier) for the user.
  2. Ask model which skills are relevant to the prompt.
  3. Load full bodies (active) + section excerpts (archive) + all system skills.
  4. Generate response (web search tool available).
  5. Reflect: model returns skill edits across system/active/archive tiers.
  6. Append turn to sessions/turns tables.
"""

from __future__ import annotations

import base64
import json
import os
import re
import secrets
import sys
import time
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone

import anthropic
from google import genai
from google.genai import errors, types

from store import SessionStore, SkillStore, UserStore

MODEL = "gemini-3.5-flash"
# Claude Haiku 4.5 (Anthropic API). Cheap/fast tier; text + vision + server-side
# web search + manual extended thinking (budget_tokens) — no image generation,
# no code execution.
CLAUDE_MODEL = "claude-haiku-4-5"
CLAUDE_SONNET_MODEL = "claude-sonnet-5"  # Sonnet 5 (Anthropic). Same v1 path as Haiku.
ALLOWED_MODELS = (
    "gemini-3.5-flash",
    "gemini-3.1-flash-lite",
    CLAUDE_MODEL,
    CLAUDE_SONNET_MODEL,
)


def provider_of(model: str) -> str:
    """Route a chat model to its SDK provider. Claude ids start with 'claude'."""
    return "anthropic" if (model or "").startswith("claude") else "gemini"

# Image generation runs on a dedicated native-image model, never user-selectable.
# The user's chat model still drives intent detection; only the pixels come from
# here. Returns inline image parts (response_modalities = TEXT + IMAGE).
IMAGE_MODEL = "gemini-3.1-flash-lite-image"

# Per-model thinking level for the main respond() call. Gemini 3 tool use
# (code execution / search) needs thinking engaged; without a thinking_config
# flash-lite mis-emits a bare function_call instead of running code. Mirror the
# levels AI Studio's generated code pairs with each model.
THINKING_LEVELS = {
    "gemini-3.5-flash": types.ThinkingLevel.MEDIUM,
    "gemini-3.1-flash-lite": types.ThinkingLevel.MEDIUM,
}

# AGENT_THINKING_LEVEL overrides the per-model default for the respond() call.
# Accepts minimal | low | medium | high (case-insensitive). Unset/invalid =>
# fall back to THINKING_LEVELS[model].
THINKING_LEVEL_NAMES = {
    "minimal": types.ThinkingLevel.MINIMAL,
    "low": types.ThinkingLevel.LOW,
    "medium": types.ThinkingLevel.MEDIUM,
    "high": types.ThinkingLevel.HIGH,
}

# Claude Haiku 4.5 supports manual extended thinking (thinking.budget_tokens).
# Claude Sonnet 5 does NOT — manual budget_tokens was replaced by always-on
# adaptive thinking + an `effort` param on Sonnet 5, and returns a 400 if sent.
# So only Haiku is wired here; the UI disables the control (shows "NA") for
# any model not in this set.
CLAUDE_THINKING_MODELS = (CLAUDE_MODEL,)

# budget_tokens per level, reusing the same minimal/low/medium/high names as
# the Gemini dropdown. max_tokens must exceed budget_tokens (see _chat_respond).
CLAUDE_THINKING_BUDGETS = {
    "minimal": 1024,
    "low": 4096,
    "medium": 10000,
    "high": 32000,
}

# Models whose "think" dropdown actually does something server-side.
THINKING_CAPABLE_MODELS = tuple(THINKING_LEVELS.keys()) + CLAUDE_THINKING_MODELS


def resolve_thinking_level(model, override=None):
    """Pick the thinking level. Precedence: explicit override (per-request, e.g.
    from the UI) -> AGENT_THINKING_LEVEL env -> per-model default -> MEDIUM."""
    for cand in (override, os.environ.get("AGENT_THINKING_LEVEL")):
        name = (cand or "").strip().lower()
        if name in THINKING_LEVEL_NAMES:
            return THINKING_LEVEL_NAMES[name]
    return THINKING_LEVELS.get(model, types.ThinkingLevel.MEDIUM)


def resolve_claude_thinking_budget(model, override=None):
    """Pick the Claude budget_tokens. Same override -> env -> default
    precedence as resolve_thinking_level; None if `model` has no manual
    thinking support (not in CLAUDE_THINKING_MODELS) or no level resolved."""
    if model not in CLAUDE_THINKING_MODELS:
        return None
    for cand in (override, os.environ.get("AGENT_THINKING_LEVEL")):
        name = (cand or "").strip().lower()
        if name in CLAUDE_THINKING_BUDGETS:
            return CLAUDE_THINKING_BUDGETS[name]
    return None

USER_ID_RE = re.compile(r"^[A-Za-z0-9_-]{1,64}$")
SESSION_ID_RE = re.compile(r"^[A-Za-z0-9_-]{1,64}$")
FRONTMATTER_RE = re.compile(r"^---\s*\n(.*?)\n---\s*\n?(.*)$", re.DOTALL)
SECTION_RE = re.compile(r"^(##\s+.*)$", re.MULTILINE)
TIERS = ("system", "active", "archive")

# Local-dev fallback path for the cookie-signing secret. Skipped entirely on
# Vercel (read-only FS); set SECRET_KEY in the environment there.
SECRET_KEY_PATH = os.path.join(os.path.dirname(__file__), ".secret_key")
IS_SERVERLESS = bool(os.environ.get("VERCEL") or os.environ.get("AWS_LAMBDA_FUNCTION_NAME"))

# Maximum number of prior turns (user+model messages) sent to `respond`.
# Avoids unbounded history growth as sessions get long. Tune via env var.
HISTORY_TURN_CAP = int(os.environ.get("AGENT_HISTORY_CAP", "20"))


def _window_history(history: list, cap: int = HISTORY_TURN_CAP) -> list:
    """Return only the last `cap` turn entries. Pure, cheap."""
    if cap <= 0 or len(history) <= cap:
        return history
    return history[-cap:]


# ---------- user context ----------

def validate_user_id(user_id: str) -> str:
    """Validate user_id against strict regex. Returns the id or raises ValueError."""
    if not isinstance(user_id, str):
        raise ValueError("user_id must be string")
    if "\x00" in user_id:
        raise ValueError("user_id contains NUL")
    if not USER_ID_RE.match(user_id):
        raise ValueError(f"invalid user_id: {user_id!r}")
    return user_id


@dataclass
class UserCtx:
    """Pure identity carrier. No filesystem state."""
    user_id: str
    model: str = MODEL
    # Optional per-request thinking level (minimal|low|medium|high). None =>
    # fall back to AGENT_THINKING_LEVEL env / per-model default in respond().
    thinking_level: str | None = None

    def __post_init__(self) -> None:
        validate_user_id(self.user_id)
        if self.model not in ALLOWED_MODELS:
            self.model = MODEL
        name = (self.thinking_level or "").strip().lower()
        self.thinking_level = name if name in THINKING_LEVEL_NAMES else None


def get_secret_key() -> bytes:
    """Bootstrap a persistent secret key for cookie signing.

    Order of preference:
      1. SECRET_KEY env var (required on Vercel).
      2. Local-dev file fallback at .secret_key (skipped on serverless).
      3. Generate-and-save (local-dev only).
    """
    env = os.environ.get("SECRET_KEY")
    if env:
        return env.encode("utf-8")
    if IS_SERVERLESS:
        raise RuntimeError(
            "SECRET_KEY env var is required in serverless deployments "
            "(no writable filesystem for the file fallback)."
        )
    if os.path.exists(SECRET_KEY_PATH):
        with open(SECRET_KEY_PATH, "rb") as f:
            return f.read().strip()
    key = secrets.token_hex(32).encode("utf-8")
    with open(SECRET_KEY_PATH, "wb") as f:
        f.write(key)
    try:
        os.chmod(SECRET_KEY_PATH, 0o600)
    except OSError:
        pass
    return key


# ---------- skill IO ----------

@dataclass
class Skill:
    name: str
    description: str
    body: str
    tier: str = "active"

    def serialize(self) -> str:
        import yaml
        fm = yaml.safe_dump(
            {"name": self.name, "description": self.description},
            sort_keys=False,
        ).strip()
        return f"---\n{fm}\n---\n\n{self.body.strip()}\n"


def load_skills(ctx: UserCtx) -> list[Skill]:
    rows = SkillStore.load_all(ctx.user_id)
    return [
        Skill(
            name=r["name"],
            description=r.get("description", ""),
            body=r.get("body", ""),
            tier=r.get("tier", "active"),
        )
        for r in rows
    ]


def write_skill(
    ctx: UserCtx, name: str, description: str, body: str, tier: str = "active"
) -> str:
    """Insert-or-update a skill. Returns the canonical name (used as identifier)."""
    if tier not in TIERS:
        tier = "active"
    return SkillStore.upsert(ctx.user_id, name, description, body, tier=tier)


def delete_skill(ctx: UserCtx, name: str) -> bool:
    """Delete a skill by name. Returns True if a row was removed."""
    return SkillStore.delete(ctx.user_id, name)


def split_sections(body: str) -> list[tuple[str, str]]:
    """Split body on ## headings. Returns [(heading, chunk_text)]."""
    parts = SECTION_RE.split(body)
    if len(parts) <= 1:
        return [("(full)", body.strip())]
    out: list[tuple[str, str]] = []
    if parts[0].strip():
        out.append(("(preamble)", parts[0].strip()))
    for i in range(1, len(parts), 2):
        heading = parts[i].strip()
        content = parts[i + 1] if i + 1 < len(parts) else ""
        out.append((heading, f"{heading}\n{content.strip()}"))
    return out


# ---------- session IO ----------

def new_session_id() -> str:
    return uuid.uuid4().hex[:16]


def new_session(ctx: UserCtx) -> str:
    sid = new_session_id()
    SessionStore.create(ctx.user_id, sid)
    return sid


def append_turn(
    ctx: UserCtx,
    session_id: str,
    role: str,
    text: str,
    usage: dict | None = None,
) -> None:
    SessionStore.append_turn(ctx.user_id, session_id, role, text, tokens=usage)


def load_session(ctx: UserCtx, session_id: str) -> list[types.Content]:
    rows = SessionStore.load_turns(ctx.user_id, session_id)
    out: list[types.Content] = []
    for r in rows:
        role = r.get("role")
        text = r.get("text", "")
        if role in ("user", "model") and text:
            out.append(types.Content(role=role, parts=[types.Part(text=text)]))
    return out


def list_sessions(ctx: UserCtx) -> list[dict]:
    """Return [{session_id, first_message, last_ts, turns, rolled_up}]."""
    return SessionStore.list_sessions(ctx.user_id)


def session_turns(ctx: UserCtx, session_id: str) -> list[dict]:
    """Return ordered chat turns [{role, text, tokens}] for replay on refresh."""
    rows = SessionStore.load_turns(ctx.user_id, session_id)
    out: list[dict] = []
    for r in rows:
        role = r.get("role")
        text = r.get("text", "")
        if role in ("user", "model") and text:
            out.append({"role": role, "text": text, "tokens": r.get("tokens")})
    return out


# ---------- LLM helpers ----------

def _client(api_key: str | None = None) -> genai.Client:
    """Build a Gemini client.

    Web deploy: each request supplies the user's own key (held only in their
    browser session) — pass it here. CLI: `api_key` is None and the SDK falls
    back to GEMINI_API_KEY / GOOGLE_API_KEY from the environment.
    """
    if api_key:
        return genai.Client(api_key=api_key)
    return genai.Client()


def _anthropic_client(api_key: str | None = None) -> anthropic.Anthropic:
    """Build an Anthropic client. Mirrors `_client`: web deploy passes the user's
    own key (held only in their browser session); CLI passes None and the SDK
    falls back to ANTHROPIC_API_KEY from the environment. The SDK retries
    429/5xx automatically (max_retries default 2), so the Claude path needs no
    hand-rolled backoff."""
    if api_key:
        return anthropic.Anthropic(api_key=api_key)
    return anthropic.Anthropic()


@dataclass
class Clients:
    """Holder so both SDKs can coexist. Image generation always needs Gemini;
    chat may route to either. Either may be None — a Claude-only web user has no
    Gemini client, and supplies no Gemini key."""
    gemini: "genai.Client | None" = None
    anthropic: "anthropic.Anthropic | None" = None


# Free-tier keys hit per-minute quotas easily: a single turn fires several
# Gemini calls (pick → archive → respond → reflect). Back off and retry on 429
# instead of letting the whole turn die.
_RETRY_BACKOFF = (2, 8, 20, 40)


def _retry_delay_from_error(e: errors.APIError) -> float | None:
    """Pull RetryInfo.retryDelay (e.g. '17s') from a 429 error, if present."""
    details = e.details or {}
    inner = details.get("error", {}) if isinstance(details, dict) else {}
    for d in inner.get("details", []) or []:
        if "RetryInfo" in str(d.get("@type", "")):
            raw = str(d.get("retryDelay", "")).rstrip("s")
            try:
                return float(raw)
            except ValueError:
                return None
    return None


def _generate(client: genai.Client, **kwargs):
    """`generate_content` with bounded backoff on 429 RESOURCE_EXHAUSTED."""
    for attempt, fallback in enumerate((*_RETRY_BACKOFF, None)):
        try:
            return client.models.generate_content(**kwargs)
        except errors.APIError as e:
            if getattr(e, "code", None) != 429 or fallback is None:
                raise
            wait = _retry_delay_from_error(e) or fallback
            print(
                f"[rate-limit] 429; retry {attempt + 1}/{len(_RETRY_BACKOFF)} "
                f"in {wait}s",
                file=sys.stderr,
            )
            time.sleep(wait)


def _usage_dict(resp) -> dict:
    u = getattr(resp, "usage_metadata", None)
    if not u:
        return {}
    return {
        "in": getattr(u, "prompt_token_count", 0) or 0,
        "out": getattr(u, "candidates_token_count", 0) or 0,
        "thoughts": getattr(u, "thoughts_token_count", 0) or 0,
        "total": getattr(u, "total_token_count", 0) or 0,
    }


def _log_usage(label: str, resp) -> dict:
    u = _usage_dict(resp)
    if not u:
        return {}
    parts = [f"in={u['in']}", f"out={u['out']}"]
    if u.get("thoughts"):
        parts.append(f"thoughts={u['thoughts']}")
    parts.append(f"total={u['total']}")
    print(f"[tokens {label}] {' '.join(parts)}")
    return u


def _anthropic_usage_dict(resp) -> dict:
    """Map an Anthropic Message's usage to the same shape as `_usage_dict`.
    Anthropic has no 'thoughts' token field; total = in + out."""
    u = getattr(resp, "usage", None)
    if not u:
        return {}
    inp = getattr(u, "input_tokens", 0) or 0
    out = getattr(u, "output_tokens", 0) or 0
    return {"in": inp, "out": out, "thoughts": 0, "total": inp + out}


def _log_anthropic_usage(label: str, resp) -> dict:
    u = _anthropic_usage_dict(resp)
    if not u:
        return {}
    print(f"[tokens {label}] in={u['in']} out={u['out']} total={u['total']}")
    return u


_JSON_FENCE_RE = re.compile(r"```(?:json)?\s*(.*?)\s*```", re.DOTALL)


def _extract_json(text: str):
    """Tolerantly pull a JSON value out of model text. Handles ```json fences and
    leading/trailing prose by falling back to the first balanced {...} or [...].
    Returns the parsed object, or None on failure. Used on the Claude path, which
    we don't assume emits bare structured output."""
    if not text:
        return None
    raw = text.strip()
    try:
        return json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        pass
    m = _JSON_FENCE_RE.search(raw)
    if m:
        try:
            return json.loads(m.group(1))
        except (json.JSONDecodeError, TypeError):
            pass
    # Fall back to the first balanced object/array in the text.
    for opener, closer in (("{", "}"), ("[", "]")):
        start = raw.find(opener)
        if start == -1:
            continue
        depth = 0
        for i in range(start, len(raw)):
            c = raw[i]
            if c == opener:
                depth += 1
            elif c == closer:
                depth -= 1
                if depth == 0:
                    try:
                        return json.loads(raw[start : i + 1])
                    except (json.JSONDecodeError, TypeError):
                        break
    return None


# Per-call max output tokens for the Claude path. Chat answers are short; this
# keeps non-streaming requests well under the SDK's HTTP timeout guard.
CLAUDE_MAX_TOKENS = 8192
# Server-side web-search tool for Haiku 4.5. The newer _20260209 (dynamic
# filtering) variant requires Opus 4.6+/Sonnet 4.6 and 400s on Haiku.
CLAUDE_WEB_SEARCH_TOOL = {"type": "web_search_20250305", "name": "web_search"}


def _chat_json(
    clients: Clients, model: str, system: str, user: str, label: str
):
    """Provider-agnostic 'return JSON' call. Returns (parsed_obj_or_None, usage).
    Gemini uses response_mime_type=application/json + json.loads; Claude uses a
    plain message + _extract_json (fence/balance-tolerant)."""
    if provider_of(model) == "anthropic":
        resp = clients.anthropic.messages.create(
            model=model,
            max_tokens=CLAUDE_MAX_TOKENS,
            system=system,
            messages=[{"role": "user", "content": user}],
        )
        usage = _log_anthropic_usage(label, resp)
        text = "".join(
            b.text for b in resp.content if getattr(b, "type", None) == "text"
        )
        return _extract_json(text), usage
    resp = _generate(
        clients.gemini,
        model=model,
        contents=user,
        config=types.GenerateContentConfig(
            system_instruction=system,
            response_mime_type="application/json",
        ),
    )
    usage = _log_usage(label, resp)
    try:
        return json.loads(resp.text), usage
    except (json.JSONDecodeError, TypeError):
        return None, usage


def _chat_respond(
    clients: Clients,
    model: str,
    sys_prompt: str,
    history: list,
    prompt: str,
    attachments: list,
    allow_code: bool,
    allow_search: bool,
    thinking_level: str | None,
    label: str,
) -> tuple[str, dict]:
    """Provider-agnostic main answer. Builds the per-provider SDK call from a
    prompt that `respond` already assembled. Returns (text, usage)."""
    if provider_of(model) == "anthropic":
        # v1 Claude path: web search (server-side) when allowed; no code
        # execution. Manual extended thinking (budget_tokens) only exists on
        # Haiku 4.5 — Sonnet 5 uses always-on adaptive thinking + an `effort`
        # param instead and 400s on budget_tokens, so it's excluded from
        # CLAUDE_THINKING_MODELS and thinking_level is a no-op for it here.
        budget = resolve_claude_thinking_budget(model, thinking_level)
        user_content = attachment_blocks(attachments or []) + [
            {"type": "text", "text": prompt}
        ]
        messages = _history_to_claude(history) + [
            {"role": "user", "content": user_content}
        ]
        tools = [CLAUDE_WEB_SEARCH_TOOL] if allow_search else []
        resp = None
        for _ in range(6):  # bound pause_turn continuations
            kwargs = dict(
                model=model,
                max_tokens=CLAUDE_MAX_TOKENS + (budget or 0),
                system=sys_prompt,
                messages=messages,
            )
            if tools:
                kwargs["tools"] = tools
            if budget:
                kwargs["thinking"] = {"type": "enabled", "budget_tokens": budget}
            resp = clients.anthropic.messages.create(**kwargs)
            if getattr(resp, "stop_reason", None) != "pause_turn":
                break
            # Server tool hit its iteration cap; re-send to resume. Do NOT add a
            # "continue" message — the trailing server_tool_use signals resume.
            messages = messages + [
                {"role": "assistant", "content": resp.content}
            ]
        usage = _log_anthropic_usage(label, resp)
        text = "\n\n".join(
            b.text
            for b in (resp.content or [])
            if getattr(b, "type", None) == "text" and getattr(b, "text", None)
        ).strip()
        return text, usage

    tools = [types.Tool(google_search=types.GoogleSearch())]
    if allow_code:
        tools.append(types.Tool(code_execution=types.ToolCodeExecution()))
    user_parts = attachment_parts(attachments or []) + [types.Part(text=prompt)]
    contents = history + [types.Content(role="user", parts=user_parts)]
    resp = _generate(
        clients.gemini,
        model=model,
        contents=contents,
        config=types.GenerateContentConfig(
            system_instruction=sys_prompt,
            tools=tools,
            thinking_config=types.ThinkingConfig(
                thinking_level=resolve_thinking_level(model, thinking_level)
            ),
        ),
    )
    if os.environ.get("AGENT_DEBUG"):
        cand = (getattr(resp, "candidates", None) or [None])[0]
        parts = getattr(getattr(cand, "content", None), "parts", None) or []
        kinds = [
            attr
            for p in parts
            for attr in (
                "text", "executable_code", "code_execution_result",
                "function_call",
            )
            if getattr(p, attr, None)
        ]
        ran_code = "executable_code" in kinds
        print(f"[respond parts] {kinds} code_executed={ran_code}")
    usage = _log_usage(label, resp)
    return _assemble_answer(resp), usage


def pick_skills(
    clients: Clients,
    prompt: str,
    skills: list[Skill],
    model: str = MODEL,
    history_snippet: str = "",
) -> list[str]:
    """Return list of skill names model deems relevant. System tier auto-included.

    `history_snippet` gives the picker recent-turn context so follow-up
    prompts ("what did I say its stack was?") still match skill descriptions.
    """
    candidates = [s for s in skills if s.tier != "system"]
    if not candidates:
        return []
    catalog = "\n".join(
        f"- [{s.tier}] {s.name}: {s.description}" for s in candidates
    )
    sys_prompt = (
        "You select relevant skills for a user prompt. "
        "Return ONLY a JSON array of skill names from the catalog. "
        "Empty array if none apply."
    )
    context = (
        f"Recent conversation:\n{history_snippet}\n\n" if history_snippet else ""
    )
    user = f"Catalog:\n{catalog}\n\n{context}User prompt:\n{prompt}\n\nJSON array:"
    names, _ = _chat_json(clients, model, sys_prompt, user, "pick")
    if not isinstance(names, list):
        return []
    valid = {s.name for s in candidates}
    return [n for n in names if n in valid]


def pick_archive_sections(
    clients: Clients,
    prompt: str,
    skill: Skill,
    model: str = MODEL,
    history_snippet: str = "",
) -> str:
    """For archive-tier skill: return only relevant ## sections."""
    sections = split_sections(skill.body)
    if len(sections) <= 1:
        return skill.body
    headings = "\n".join(f"- {i}: {h}" for i, (h, _) in enumerate(sections))
    sys_prompt = (
        "Pick relevant sections from archived skill for the prompt. "
        "Return ONLY a JSON array of section indices (integers)."
    )
    context = (
        f"Recent conversation:\n{history_snippet}\n\n" if history_snippet else ""
    )
    user = (
        f"Skill: {skill.name}\nDescription: {skill.description}\n"
        f"Sections:\n{headings}\n\n{context}Prompt:\n{prompt}\n\nJSON array of indices:"
    )
    idxs, _ = _chat_json(clients, model, sys_prompt, user, f"archive:{skill.name}")
    if not isinstance(idxs, list):
        return ""
    chunks = [
        sections[i][1]
        for i in idxs
        if isinstance(i, int) and 0 <= i < len(sections)
    ]
    return "\n\n".join(chunks)


def _assemble_answer(resp) -> str:
    """Build a readable markdown answer from a (possibly multi-part) response.

    With the code_execution tool, a response interleaves text, executable_code,
    and code_execution_result parts. `resp.text` drops the latter two (and
    warns), so we walk parts in order and fence code + output. Falls back to
    `resp.text` when there are no structured parts."""
    cand = (getattr(resp, "candidates", None) or [None])[0]
    content = getattr(cand, "content", None) if cand else None
    parts = getattr(content, "parts", None) if content else None
    if not parts:
        return resp.text or ""
    out: list[str] = []
    saw_function_call = False
    for part in parts:
        text = getattr(part, "text", None)
        if text:
            out.append(text)
        code = getattr(part, "executable_code", None)
        if code is not None and getattr(code, "code", None):
            out.append(f"```python\n{code.code}\n```")
        result = getattr(part, "code_execution_result", None)
        if result is not None and getattr(result, "output", None):
            out.append(f"```\n{result.output}\n```")
        # A bare function_call (no matching callable) would otherwise leave the
        # answer empty. Surface it instead of silently dropping the turn.
        fc = getattr(part, "function_call", None)
        if fc is not None and getattr(fc, "name", None):
            saw_function_call = True
    assembled = "\n\n".join(out).strip()
    if assembled:
        return assembled
    if saw_function_call:
        return (
            "(The model returned an unhandled tool call and no text - the tool "
            "invocation did not resolve to a result this turn. Try rephrasing, "
            "or raise the model's thinking level.)"
        )
    return resp.text or ""


def respond(
    clients: Clients,
    prompt: str,
    system_skills: list[Skill],
    active_loaded: list[Skill],
    archive_excerpts: list[tuple[Skill, str]],
    history: list[types.Content],
    model: str = MODEL,
    allow_code_execution: bool = False,
    thinking_level: str | None = None,
    attachments: list[Attachment] | None = None,
) -> tuple[str, dict]:
    """Main answer. Has web-search grounding (Gemini google_search / Anthropic
    server-side web_search) and, on Gemini, optional code execution. Attachments
    ride as extra parts/blocks on the user message. Returns (text, usage).

    Prompt-building lives here; the per-provider SDK call is delegated to
    `_chat_respond`. Code execution is Gemini-only in v1 — `allow_code_execution`
    is ignored on the Claude path."""
    is_anthropic = provider_of(model) == "anthropic"
    allow_code = allow_code_execution and not is_anthropic
    sys_block = "\n\n".join(
        f"## [system] {s.name}\n{s.body.strip()}" for s in system_skills
    )
    active_block = "\n\n".join(
        f"## [active] {s.name}\n{s.description}\n\n{s.body.strip()}"
        for s in active_loaded
    )
    archive_block = "\n\n".join(
        f"## [archive excerpt] {s.name}\n{excerpt}"
        for s, excerpt in archive_excerpts
        if excerpt
    )
    loaded_text = "\n\n".join(
        b for b in (sys_block, active_block, archive_block) if b
    )
    search_tool = "web_search" if is_anthropic else "google_search"
    sys_prompt = (
        "You are a helpful agent with a skill-memory system. "
        "Use the loaded skills below as long-term memory. "
        f"Use {search_tool} when current information is needed.\n\n"
        f"LOADED SKILLS:\n{loaded_text or '(none)'}"
    )
    if allow_code:
        sys_prompt += (
            "\n\nYou may run Python via the code execution tool for "
            "calculation, data manipulation, or anything better solved by "
            "executing code than by reasoning in prose."
        )
    if attachments:
        sys_prompt += (
            "\n\nThe user attached files to this message (documents, images, "
            "or spreadsheet data). They are included with the message — read "
            "them and answer based on their actual content."
        )
    return _chat_respond(
        clients,
        model,
        sys_prompt,
        history,
        prompt,
        attachments,
        allow_code=allow_code,
        allow_search=True,
        thinking_level=thinking_level,
        label="respond",
    )


# ---------- attachments (session-only uploads) ----------
#
# Files ride along with a chat request as extra Gemini parts. They are NEVER
# persisted server-side — only a text marker ("[attached: …]") lands in the
# turns table, so reflect/session-summary still capture WHAT was discussed.
# The browser keeps the bytes for the life of the tab session and re-sends
# them each turn so follow-up questions keep working.

ATTACH_MAX_FILES = 5
ATTACH_MAX_FILE_BYTES = 8 * 1024 * 1024    # per file
ATTACH_MAX_TOTAL_BYTES = 16 * 1024 * 1024  # per request (Gemini inline ~20MB)
ATTACH_SHEET_ROW_CAP = 500                 # rows per sheet sent to the model

ATTACH_TEXT_MIMES = {
    "application/json",
    "application/xml",
    "application/x-yaml",
    "application/yaml",
    "text/csv",
    "text/tab-separated-values",
}

ATTACH_XLSX_MIME = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


@dataclass
class Attachment:
    """One validated upload. Either raw bytes (pdf/image — Gemini reads them
    natively) or extracted text (text docs, spreadsheets)."""
    name: str
    mime: str
    data: bytes            # empty when `text` carries the content
    text: str | None
    size: int              # original upload size, for display


def _xlsx_to_text(data: bytes) -> str:
    """Render an .xlsx workbook as CSV text, one section per sheet."""
    import csv as _csv
    import io

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out: list[str] = []
    try:
        for ws in wb.worksheets:
            buf = io.StringIO()
            writer = _csv.writer(buf)
            truncated = False
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= ATTACH_SHEET_ROW_CAP:
                    truncated = True
                    break
                writer.writerow(["" if c is None else c for c in row])
            section = f"### Sheet: {ws.title}\n{buf.getvalue().strip()}"
            if truncated:
                section += f"\n(truncated at {ATTACH_SHEET_ROW_CAP} rows)"
            out.append(section)
    finally:
        wb.close()
    return "\n\n".join(out)


def prepare_attachments(raw: list[dict]) -> list[Attachment]:
    """Decode + validate uploads ({name, mime, data(b64)} dicts).

    Raises ValueError with a user-facing message on anything off-spec."""
    if len(raw) > ATTACH_MAX_FILES:
        raise ValueError(f"too many attachments (max {ATTACH_MAX_FILES})")
    total = 0
    out: list[Attachment] = []
    for item in raw:
        name = os.path.basename(str(item.get("name") or "file")).strip()[:120] or "file"
        mime = (str(item.get("mime") or "").strip().lower()
                or "application/octet-stream")
        try:
            data = base64.b64decode(item.get("data") or "", validate=True)
        except (ValueError, TypeError):
            raise ValueError(f"{name}: invalid base64 payload")
        if not data:
            raise ValueError(f"{name}: empty file")
        if len(data) > ATTACH_MAX_FILE_BYTES:
            raise ValueError(
                f"{name}: too large (max {ATTACH_MAX_FILE_BYTES // (1024 * 1024)}MB per file)"
            )
        total += len(data)
        if total > ATTACH_MAX_TOTAL_BYTES:
            raise ValueError(
                f"attachments too large together (max {ATTACH_MAX_TOTAL_BYTES // (1024 * 1024)}MB)"
            )
        size = len(data)
        if mime == "application/pdf" or mime.startswith("image/"):
            out.append(Attachment(name, mime, data, None, size))
        elif mime == ATTACH_XLSX_MIME or name.lower().endswith(".xlsx"):
            try:
                text = _xlsx_to_text(data)
            except Exception as e:
                raise ValueError(f"{name}: could not read spreadsheet ({e})")
            out.append(Attachment(name, "text/csv", b"", text, size))
        elif mime.startswith("text/") or mime in ATTACH_TEXT_MIMES:
            out.append(
                Attachment(name, mime, b"", data.decode("utf-8", errors="replace"), size)
            )
        else:
            # Unknown mime: accept if it decodes cleanly as UTF-8 text.
            try:
                text = data.decode("utf-8")
            except UnicodeDecodeError:
                raise ValueError(
                    f"{name}: unsupported file type ({mime}). "
                    "Supported: PDF, images, spreadsheets (.xlsx/.csv), text docs."
                )
            out.append(Attachment(name, "text/plain", b"", text, size))
    return out


def attachment_parts(attachments: list[Attachment]) -> list[types.Part]:
    """Gemini parts for the user message: native bytes for pdf/image, fenced
    text blocks for everything else."""
    parts: list[types.Part] = []
    for a in attachments:
        if a.text is not None:
            parts.append(types.Part(
                text=(
                    f"--- ATTACHED FILE: {a.name} ({a.mime}) ---\n"
                    f"{a.text}\n--- END FILE: {a.name} ---"
                )
            ))
        else:
            parts.append(types.Part.from_bytes(data=a.data, mime_type=a.mime))
    return parts


def attachment_blocks(attachments: list[Attachment]) -> list[dict]:
    """Anthropic content blocks for the user message: a text block (with the same
    marker as the Gemini path) for extracted text/spreadsheets, a base64 image
    block for images, and a base64 document block for PDFs (placed before any
    text by the caller). On the Claude path image-only uploads arrive here as
    vision inputs because the image-generation fast-path is skipped."""
    blocks: list[dict] = []
    for a in attachments:
        if a.text is not None:
            blocks.append({
                "type": "text",
                "text": (
                    f"--- ATTACHED FILE: {a.name} ({a.mime}) ---\n"
                    f"{a.text}\n--- END FILE: {a.name} ---"
                ),
            })
        elif a.mime == "application/pdf":
            blocks.append({
                "type": "document",
                "source": {
                    "type": "base64",
                    "media_type": "application/pdf",
                    "data": base64.b64encode(a.data).decode("ascii"),
                },
            })
        else:  # image/*
            blocks.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": a.mime,
                    "data": base64.b64encode(a.data).decode("ascii"),
                },
            })
    return blocks


def _history_to_claude(history: list[types.Content]) -> list[dict]:
    """Convert windowed Gemini history (list[types.Content]) to Anthropic message
    dicts. Gemini role 'model' maps to 'assistant'; part texts are joined."""
    out: list[dict] = []
    for c in history:
        role = "assistant" if c.role == "model" else "user"
        text = "".join(p.text or "" for p in (c.parts or []))
        if text:
            out.append({"role": role, "content": text})
    return out


def attachment_manifest(attachments: list[Attachment]) -> str:
    """Text markers persisted to the turns table in place of the bytes."""
    return "\n".join(
        f"[attached: {a.name} ({a.mime}, {max(1, a.size // 1024)} KB)]"
        for a in attachments
    )


# ---------- image generation ----------

def _image_gen_enabled() -> bool:
    """Image generation toggle. On by default; set AGENT_IMAGE_GEN=0 to disable
    (skips the per-turn intent classifier call to conserve quota)."""
    return os.environ.get("AGENT_IMAGE_GEN", "1").strip().lower() not in (
        "0", "false", "no", "off", ""
    )


_IMAGE_INTENT_INSTRUCTIONS = (
    "Decide whether the user is asking you to CREATE / GENERATE / DRAW / PAINT a "
    "NEW image (picture, illustration, logo, diagram-as-art, etc.). "
    "Editing or analysing an existing image, or merely discussing images, is NOT "
    "image generation. "
    'Return ONLY JSON: {"image": true|false, "prompt": "<standalone, vivid image '
    'description>"}. When image is false, prompt may be empty.'
)

_IMAGE_INTENT_INSTRUCTIONS_WITH_REFS = (
    "The user has attached reference image(s) to this message. Decide whether "
    "they are asking you to CREATE / GENERATE a NEW image FROM or BASED ON those "
    "reference image(s) — e.g. restyle, transform, reimagine, redraw the subject "
    "in a different style or setting. That IS image generation. "
    "Merely analysing, describing, or answering questions about the attached "
    "image(s) is NOT image generation. "
    'Return ONLY JSON: {"image": true|false, "prompt": "<standalone instruction '
    "for an image model that will also receive the reference image(s); describe "
    'the transformation, do not re-describe the reference content>"}. '
    "When image is false, prompt may be empty."
)


def detect_image_intent(
    clients: Clients, prompt: str, model: str = MODEL,
    has_reference_images: bool = False,
) -> dict:
    """Cheap classifier: does the user want a NEW image? Returns
    {"image": bool, "prompt": str}. Failure-safe — returns image=False on any
    parse/transport hiccup so a turn never dies in the gate. Only reached on the
    Gemini path (the fast-path is gated to Gemini), but routes through the shared
    adapter for consistency."""
    system = (
        _IMAGE_INTENT_INSTRUCTIONS_WITH_REFS
        if has_reference_images
        else _IMAGE_INTENT_INSTRUCTIONS
    )
    data, _ = _chat_json(
        clients, model, system, f"User prompt:\n{prompt}\n\nJSON:", "image-intent"
    )
    if isinstance(data, dict) and data.get("image"):
        return {"image": True, "prompt": (data.get("prompt") or prompt).strip()}
    return {"image": False, "prompt": ""}


def generate_image(
    clients: Clients, image_prompt: str,
    reference_images: list[Attachment] | None = None,
) -> tuple[bytes | None, str, str]:
    """Call the native-image model. Returns (image_bytes|None, mime, caption_text).
    image_bytes is None when the model declined to emit an image.
    `reference_images` (image attachments) ride along as input parts for
    image-to-image generation."""
    contents: str | list[types.Part] = image_prompt
    if reference_images:
        contents = attachment_parts(reference_images) + [
            types.Part(text=image_prompt)
        ]
    resp = _generate(
        clients.gemini,
        model=IMAGE_MODEL,
        contents=contents,
        config=types.GenerateContentConfig(
            response_modalities=["TEXT", "IMAGE"],
        ),
    )
    _log_usage("image", resp)
    cand = (getattr(resp, "candidates", None) or [None])[0]
    content = getattr(cand, "content", None) if cand else None
    parts = getattr(content, "parts", None) if content else None
    data: bytes | None = None
    mime = "image/png"
    captions: list[str] = []
    for part in parts or []:
        inline = getattr(part, "inline_data", None)
        if inline is not None and getattr(inline, "data", None):
            data = inline.data
            mime = getattr(inline, "mime_type", None) or mime
        text = getattr(part, "text", None)
        if text:
            captions.append(text)
    return data, mime, "\n\n".join(captions).strip()


def _store_image_blob(data: bytes, mime: str) -> str | None:
    """Best-effort upload to Vercel Blob (the 'cheap storage' path). Returns a
    public URL or None. Active ONLY when BLOB_READ_WRITE_TOKEN is set; any
    failure degrades silently to inline-only (image still shown, just not
    persisted across reloads)."""
    token = os.environ.get("BLOB_READ_WRITE_TOKEN", "").strip()
    if not token:
        return None
    import urllib.error
    import urllib.request

    ext = (mime.split("/", 1)[-1] or "png").split("+", 1)[0]
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    pathname = f"generated-images/{stamp}-{secrets.token_hex(4)}.{ext}"
    try:
        req = urllib.request.Request(
            f"https://blob.vercel-storage.com/{pathname}",
            data=data,
            method="PUT",
            headers={
                "authorization": f"Bearer {token}",
                "x-api-version": "7",
                "x-content-type": mime,
                "content-type": mime,
            },
        )
        with urllib.request.urlopen(req, timeout=20) as r:
            payload = json.loads(r.read().decode("utf-8"))
        return payload.get("url")
    except (urllib.error.URLError, OSError, ValueError, TypeError) as e:
        print(f"[blob] upload failed, inline-only: {e}", file=sys.stderr)
        return None


def _remember_image(
    ctx: UserCtx, skills: list[Skill], image_prompt: str, url: str | None
) -> str:
    """Persist the FACT that an image was generated (prompt only — image bytes
    are never stored in the DB). Appends to an `image-generations` active skill,
    recording the public URL too when cheap storage produced one."""
    name = "image-generations"
    existing = next((s for s in skills if s.name == name), None)
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    line = f'- [{stamp}] "{image_prompt}"'
    if url:
        line += f" → {url}"
    if existing:
        body = existing.body.rstrip() + "\n" + line
        desc = existing.description
        op = "update"
    else:
        body = (
            "Images the user has asked me to generate. Prompts are kept here; "
            "the image bytes themselves are not stored unless cheap object "
            "storage is configured (then a URL is recorded).\n\n" + line
        )
        desc = "Log of images generated for the user, with their prompts."
        op = "create"
    write_skill(ctx, name, desc, body, tier="active")
    return f"{op}[active] -> {name}"


EDIT_INSTRUCTIONS = """\
You are a memory writer. Aggressively persist anything from the exchange that could
inform FUTURE conversations. Skills are markdown docs stored with a YAML
frontmatter (name, description) and a markdown body. Skills live in one of three tiers:

- "system": ALWAYS loaded into context. Reserved for core identity/rules the agent
  must obey every turn (e.g. user-profile, response-style, hard rules).
- "active" (default): Loaded only when frontmatter matches the prompt. Use for most
  facts, projects, preferences, references.
- "archive": Loaded selectively by section. Use when content is large (long notes,
  research, multi-topic reference). Structure body with ## headings — each becomes
  retrievable on its own.

WRITE A SKILL whenever the user reveals ANY of:
- Their name, role, employer, team, location, or background  → system tier
- Strong preferences for how to respond, nicknames assigned to you  → system tier
- Tools, frameworks, languages they use  → active
- Projects they work on  → active
- Rules ("always X", "never Y")  → system tier
- People, pets, customers, systems referenced  → active
- External resources (URLs, dashboards, docs)  → active
- Bulky research notes, long-form knowledge  → archive (with ## sections)

PREFER MANY SMALL SKILLS over one big one for active tier. Examples of good names:
"user-profile", "response-style", "tech-stack", "project-<name>", "rule-<topic>",
"user-pets", "interest-<thing>".

DEDUPLICATION (CRITICAL):
- BEFORE deciding op="create", scan the "Existing skills" inventory below.
- If ANY existing skill name or description covers the same topic — even partly —
  you MUST use op="update" with that exact existing name. Do NOT invent a new name
  for the same topic. Do NOT write parallel skills like "user-pets" and
  "pets-info" for the same facts.
- For op="update", produce the MERGED body (existing body + new info, no duplicate
  bullets). The "Existing bodies" section is provided so you can do this merge.
- For archive updates, preserve existing ## sections; append new ## sections only
  for genuinely new topics.

BODY FORMAT (CRITICAL):
- The "body" field is markdown ONLY. It must NOT contain a YAML frontmatter block.
- Do NOT start the body with "---". Do NOT include "name:" or "description:" lines
  inside the body. The frontmatter is constructed from the "name" and "description"
  fields you return separately.
- Bad body (DO NOT DO THIS):
    ---
    name: identity
    description: ...
    ---
    - bullet
- Good body:
    - bullet

Return ONLY JSON, no prose:
{"edits": [{"op": "create"|"update", "tier": "system"|"active"|"archive",
  "name": "...", "description": "...", "body": "..."}]}

Empty edits array ONLY if the exchange is pure small talk with zero new facts.
"""


def _strip_leading_frontmatter(body: str) -> str:
    """Remove a YAML frontmatter block if the model erroneously embedded one
    at the top of the body field."""
    if not body:
        return body
    stripped = body.lstrip()
    if not stripped.startswith("---"):
        return body
    m = FRONTMATTER_RE.match(stripped)
    if m:
        return m.group(2).lstrip()
    return body


def reflect_and_edit(
    clients: Clients,
    prompt: str,
    answer: str,
    all_skills: list[Skill],
    scoped_skills: list[Skill],
    model: str = MODEL,
) -> list[dict]:
    """Reflect on the exchange and emit skill edits.

    Cost-shaped inputs:
      - `all_skills`: every skill (any tier) — used only to build the *cheap*
        name+description inventory so the model can avoid dup-by-name across
        the whole catalog.
      - `scoped_skills`: skills whose full bodies we will pay for. Caller MUST
        pre-filter to the skills relevant to this turn (system + picked active).
        Archive bodies are explicitly excluded — archive is cold storage,
        modified only by the consolidation pass.
    """
    inventory = "\n".join(
        f"- [{s.tier}] {s.name}: {s.description}" for s in all_skills
    ) or "(no skills yet)"
    body_skills = [s for s in scoped_skills if s.tier != "archive"]
    existing_bodies = "\n\n".join(
        f"### [{s.tier}] {s.name}\n{s.body.strip()}" for s in body_skills
    )
    total_body_chars = sum(len(s.body) for s in all_skills)
    scoped_body_chars = sum(len(s.body) for s in body_skills)
    print(
        f"[reflect scope] catalog={len(all_skills)} "
        f"scoped_bodies={len(body_skills)} "
        f"total_body_chars={total_body_chars} "
        f"scoped_body_chars={scoped_body_chars}"
    )
    user = (
        f"Existing skills:\n{inventory}\n\n"
        f"Existing bodies (for update merges):\n{existing_bodies or '(none)'}\n\n"
        f"User prompt:\n{prompt}\n\nAssistant answer:\n{answer}\n\n"
        "Return edits JSON:"
    )
    data, _ = _chat_json(clients, model, EDIT_INSTRUCTIONS, user, "reflect")
    if os.environ.get("AGENT_DEBUG"):
        print(f"[reflect parsed]: {str(data)[:500]}")
    if isinstance(data, dict):
        edits = data.get("edits", [])
        return edits if isinstance(edits, list) else []
    return []


SESSION_INSTRUCTIONS = """\
Summarize this chat session as a skill memory for future sessions.
Capture: topics discussed, decisions made, user state/mood, open threads, anything
worth remembering next time. 3-7 tight bullets. No filler.

Return ONLY JSON:
{"description": "one-line hook for matching future prompts", "body": "markdown bullets"}

If the session was trivial (one or two turns of small talk, no substance), return:
{"description": "", "body": ""}
"""


def summarize_session_to_skill(
    clients: Clients, ctx: UserCtx, session_id: str
) -> str | None:
    """Roll a chat into a `session-<ts>` active skill. Returns the skill name
    or None if the session was trivial or already rolled up. Routes to the
    provider of ctx.model, so a Claude-only user's session summarizes on
    Anthropic."""
    if SessionStore.is_rolled_up(ctx.user_id, session_id):
        return None
    history = load_session(ctx, session_id)
    if len(history) < 2:
        return None
    transcript = "\n\n".join(
        f"{c.role.upper()}: {''.join(p.text or '' for p in c.parts)}"
        for c in history
    )
    data, _ = _chat_json(
        clients,
        ctx.model,
        SESSION_INSTRUCTIONS,
        f"Transcript:\n{transcript}\n\nReturn JSON:",
        "session",
    )
    if isinstance(data, dict) and data.get("description") and data.get("body"):
        # Reuse the prior skill name when re-summarizing a returned-to chat,
        # so the updated summary overwrites it instead of leaving a stale
        # duplicate. Fresh sessions get a new timestamped name.
        name = SessionStore.rollup_skill(ctx.user_id, session_id)
        if not name:
            stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            name = f"session-{stamp}"
        write_skill(ctx, name, data["description"], data["body"], tier="active")
        SessionStore.mark_rolled_up(ctx.user_id, session_id, name)
        return name
    return None


# ---------- consolidation (offline pass) ----------

def consolidate(ctx: UserCtx, dry_run: bool = False) -> dict:
    """Fold every active `session-*` skill into one new archive skill and delete
    the originals. Idempotent: a no-op when there are no `session-*` skills
    left. Runs in a single transaction so a crash mid-fold leaves both the
    new archive skill and the old session skills present, never neither.
    """
    skills = load_skills(ctx)
    session_skills = [
        s for s in skills
        if s.tier == "active" and s.name.startswith("session-")
    ]

    summary = {
        "input_total": len(skills),
        "input_sessions": len(session_skills),
        "input_keep": len(skills) - len(session_skills),
        "merged_archive_skill": None,
        "deleted_session_skills": [],
        "dry_run": dry_run,
    }

    if not session_skills:
        summary["reason"] = "no session-* skills to consolidate"
        return summary

    # Build one archive skill of all session bodies, one `##` section each.
    # Heading carries the session's one-line description: pick_archive_sections
    # shows the model headings only, and a bare timestamp is unmatchable.
    sections = []
    for s in sorted(session_skills, key=lambda x: x.name):
        heading = f"{s.name} — {s.description}" if s.description else s.name
        sections.append(f"## {heading}\n\n{s.body.strip()}")
    body = "\n\n".join(sections)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    archive_name = f"sessions-archive-{stamp}"
    desc = (
        f"Consolidated {len(session_skills)} session rollups "
        f"(merged {datetime.now().date().isoformat()})."
    )

    if dry_run:
        summary["merged_archive_skill"] = archive_name
        summary["deleted_session_skills"] = [s.name for s in session_skills]
        return summary

    # Write archive skill, then delete originals. Both happen against the same
    # connection so a transactional wrapper at the caller can roll the pair
    # back together if needed; SkillStore.upsert/delete are individually atomic.
    write_skill(ctx, archive_name, desc, body, tier="archive")
    deleted: list[str] = []
    for s in session_skills:
        if delete_skill(ctx, s.name):
            deleted.append(s.name)

    summary["merged_archive_skill"] = archive_name
    summary["deleted_session_skills"] = deleted
    return summary


def apply_edits(ctx: UserCtx, edits: list[dict]) -> list[str]:
    applied: list[str] = []
    for e in edits:
        name = e.get("name")
        desc = e.get("description", "")
        body = e.get("body", "")
        tier = e.get("tier", "active")
        if not name or not body:
            continue
        body = _strip_leading_frontmatter(body)
        if not body.strip():
            continue
        canonical = write_skill(ctx, name, desc, body, tier=tier)
        applied.append(f"{e.get('op', 'write')}[{tier}] -> {canonical}")
    return applied


# ---------- shared per-turn driver (used by CLI + server) ----------

@dataclass
class TurnResult:
    answer: str
    loaded_system: list[str]
    loaded_active: list[str]
    loaded_archive: list[str]
    edits_applied: list[str]
    tokens: dict
    image: dict | None = None


def run_turn_events(
    clients: Clients,
    ctx: UserCtx,
    session_id: str,
    prompt: str,
    allow_code_execution: bool = False,
    attachments: list[Attachment] | None = None,
):
    """Generator. Yields {stage, msg, ...} dicts. Final event has stage='done'.

    `allow_code_execution` enables Gemini's code execution tool in `respond`
    (ignored on the Claude path). CLI passes True; the server leaves it False to
    keep the web path text-only.

    `attachments` are session-only uploads: their bytes go to the model for this
    turn but only a text marker is persisted (see attachment_manifest).
    """
    yield {"stage": "load", "msg": "Reading skill catalog…"}
    skills = load_skills(ctx)
    system_skills = [s for s in skills if s.tier == "system"]

    # The text we persist + reflect on: prompt plus markers for any uploads, so
    # memory captures that (and which) files were discussed without the bytes.
    persisted_prompt = prompt
    if attachments:
        persisted_prompt = f"{prompt}\n\n{attachment_manifest(attachments)}"

    # Image-generation fast path: if the user is asking for a NEW picture, route
    # to the native-image model and return early — no skill picking / respond /
    # reflect. The image rides in the final `done` event (and a live `image`
    # event) but is NOT written to the turns table; only a prompt note persists.
    # Image-only attachments stay eligible: they become reference inputs for
    # image-to-image ("redraw me as X"). Any non-image attachment (pdf, text,
    # spreadsheet) skips the path — those turns are about analysing the uploads.
    # Image generation is Gemini-only: skip the fast-path entirely on the Claude
    # path (and whenever there's no Gemini client), so a Claude user needs no
    # Gemini key and a "draw …" prompt just gets a text reply.
    image_refs = [
        a for a in (attachments or []) if a.mime.startswith("image/")
    ]
    images_only = len(image_refs) == len(attachments or [])
    gemini_image_ok = (
        provider_of(ctx.model) == "gemini" and clients.gemini is not None
    )
    if _image_gen_enabled() and images_only and gemini_image_ok:
        yield {"stage": "image_intent", "msg": "Checking for an image request…"}
        intent = detect_image_intent(
            clients, prompt, model=ctx.model,
            has_reference_images=bool(image_refs),
        )
        if intent.get("image"):
            img_prompt = intent.get("prompt") or prompt
            yield {"stage": "image_gen", "msg": "Generating image…"}
            data, mime, caption = generate_image(
                clients, img_prompt, reference_images=image_refs or None
            )
            loaded_summary = {
                "system": [s.name for s in system_skills],
                "active": [],
                "archive": [],
            }
            if not data:
                answer = caption or (
                    "I couldn't generate an image for that prompt — the model "
                    "returned no image. Try rephrasing."
                )
                append_turn(ctx, session_id, "user", persisted_prompt)
                append_turn(ctx, session_id, "model", answer)
                yield {
                    "stage": "done",
                    "answer": answer,
                    "loaded": loaded_summary,
                    "edits": [],
                    "tokens": {},
                    "meta": {},
                }
                return

            url = _store_image_blob(data, mime)
            answer = caption or f'Here is the image you asked for: "{img_prompt}".'
            # Persist text only. With cheap storage, embed the URL as markdown so
            # the image reappears on reload; otherwise the bytes are transient.
            persisted = f"{answer}\n\n![generated image]({url})" if url else answer
            append_turn(ctx, session_id, "user", persisted_prompt)
            append_turn(ctx, session_id, "model", persisted)

            edits = [_remember_image(ctx, skills, img_prompt, url)]
            loaded_summary["active"] = ["image-generations"]

            image_payload = {
                "b64": base64.b64encode(data).decode("ascii"),
                "mime": mime,
                "prompt": img_prompt,
                "url": url,
            }
            yield {
                "stage": "image",
                "msg": "Image ready.",
                "image": image_payload,
            }
            yield {
                "stage": "done",
                "answer": answer,
                "image": image_payload,
                "loaded": loaded_summary,
                "edits": edits,
                "tokens": {},
                "meta": {},
            }
            return

    yield {
        "stage": "pick",
        "msg": f"Picking relevant skills from {len(skills) - len(system_skills)} options…",
    }
    full_history = load_session(ctx, session_id)
    recent = "\n".join(
        f"{c.role}: {''.join(p.text or '' for p in c.parts)[:500]}"
        for c in full_history[-4:]
    )
    chosen_names = pick_skills(
        clients, persisted_prompt, skills, model=ctx.model, history_snippet=recent
    )
    picked = [s for s in skills if s.name in chosen_names]
    active_loaded = [s for s in picked if s.tier == "active"]
    archive_picked = [s for s in picked if s.tier == "archive"]

    archive_excerpts: list[tuple[Skill, str]] = []
    for s in archive_picked:
        yield {"stage": "archive", "msg": f"Retrieving sections from archive: {s.name}"}
        excerpt = pick_archive_sections(
            clients, prompt, s, model=ctx.model, history_snippet=recent
        )
        archive_excerpts.append((s, excerpt))

    loaded_summary = {
        "system": [s.name for s in system_skills],
        "active": [s.name for s in active_loaded],
        "archive": [s.name for s, _ in archive_excerpts],
    }
    yield {
        "stage": "respond",
        "msg": "Thinking…",
        "loaded": loaded_summary,
    }
    history = _window_history(full_history)
    if len(history) < len(full_history):
        print(
            f"[history window] full={len(full_history)} sent={len(history)}"
        )
    answer, usage = respond(
        clients,
        prompt,
        system_skills,
        active_loaded,
        archive_excerpts,
        history,
        model=ctx.model,
        allow_code_execution=allow_code_execution,
        thinking_level=ctx.thinking_level,
        attachments=attachments,
    )

    yield {"stage": "persist", "msg": "Saving turn to session log…"}
    append_turn(ctx, session_id, "user", persisted_prompt)
    append_turn(ctx, session_id, "model", answer, usage=usage)

    scoped_for_reflect = system_skills + active_loaded
    if _reflect_enabled():
        yield {"stage": "reflect", "msg": "Reflecting on what to remember…"}
        edits = reflect_and_edit(
            clients,
            persisted_prompt,
            answer,
            all_skills=skills,
            scoped_skills=scoped_for_reflect,
            model=ctx.model,
        )
        applied = apply_edits(ctx, edits)
    else:
        # Skip the reflect call to conserve per-minute quota (free-tier keys).
        applied = []

    meta = {
        "catalog": len(skills),
        "scoped_bodies": len(
            [s for s in scoped_for_reflect if s.tier != "archive"]
        ),
        "total_body_chars": sum(len(s.body) for s in skills),
        "scoped_body_chars": sum(
            len(s.body)
            for s in scoped_for_reflect
            if s.tier != "archive"
        ),
        "history_full": len(full_history),
        "history_sent": len(history),
    }

    yield {
        "stage": "done",
        "answer": answer,
        "loaded": loaded_summary,
        "edits": applied,
        "tokens": usage,
        "meta": meta,
    }


def run_turn(
    clients: Clients,
    ctx: UserCtx,
    session_id: str,
    prompt: str,
    allow_code_execution: bool = False,
    attachments: list[Attachment] | None = None,
) -> TurnResult:
    """Drains run_turn_events. Kept for CLI compatibility."""
    final: dict | None = None
    for ev in run_turn_events(
        clients, ctx, session_id, prompt,
        allow_code_execution=allow_code_execution,
        attachments=attachments,
    ):
        if ev.get("stage") == "done":
            final = ev
    if final is None:
        raise RuntimeError("run_turn_events did not emit 'done'")
    return TurnResult(
        answer=final["answer"],
        loaded_system=final["loaded"]["system"],
        loaded_active=final["loaded"]["active"],
        loaded_archive=final["loaded"]["archive"],
        edits_applied=final["edits"],
        tokens=final["tokens"],
        image=final.get("image"),
    )


# ---------- REPL ----------

def _code_exec_enabled() -> bool:
    """CLI code execution toggle. On by default; set AGENT_CODE_EXEC=0 to disable."""
    return os.environ.get("AGENT_CODE_EXEC", "1").strip().lower() not in (
        "0", "false", "no", "off", ""
    )


def _reflect_enabled() -> bool:
    """Per-turn memory reflection toggle. On by default; set AGENT_REFLECT=0 to
    skip the reflect call and save one Gemini request per turn (free-tier quota).
    Session-end summarization still runs and persists memory."""
    return os.environ.get("AGENT_REFLECT", "1").strip().lower() not in (
        "0", "false", "no", "off", ""
    )


def main() -> None:
    if "--consolidate" in sys.argv:
        _run_consolidate_cli()
        return
    if not os.environ.get("DATABASE_URL"):
        print("Set DATABASE_URL (Postgres) before running.", file=sys.stderr)
        sys.exit(1)

    user_id = os.environ.get("AGENT_USER_ID") or "default"
    try:
        validate_user_id(user_id)
    except ValueError as e:
        print(f"Invalid AGENT_USER_ID: {e}", file=sys.stderr)
        sys.exit(1)

    # AGENT_MODEL lets the CLI exercise either provider; invalid values fall back
    # to MODEL via UserCtx.__post_init__. Only the chosen provider's key is
    # required at startup.
    ctx = UserCtx(user_id=user_id, model=os.environ.get("AGENT_MODEL") or MODEL)
    has_gemini = bool(
        os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
    )
    has_anthropic = bool(os.environ.get("ANTHROPIC_API_KEY"))
    if provider_of(ctx.model) == "anthropic":
        if not has_anthropic:
            print("Set ANTHROPIC_API_KEY for the selected model.", file=sys.stderr)
            sys.exit(1)
    elif not has_gemini:
        print("Set GEMINI_API_KEY or GOOGLE_API_KEY.", file=sys.stderr)
        sys.exit(1)

    clients = Clients(
        gemini=_client() if has_gemini else None,
        anthropic=_anthropic_client() if has_anthropic else None,
    )
    session_id = new_session(ctx)
    code_exec = _code_exec_enabled()
    print(f"Agent ready ({ctx.model}). User: {user_id}. Session: {session_id}")
    print(f"Code execution: {'on' if code_exec else 'off'} (AGENT_CODE_EXEC)")
    print("Type prompt. /attach <path> to add a file (pdf/image/xlsx/csv/text), "
          "/detach to clear. Ctrl-C to exit.\n")

    attachments: list[Attachment] = []
    while True:
        try:
            prompt = input("you> ").strip()
        except (EOFError, KeyboardInterrupt):
            print()
            break
        if not prompt:
            continue

        if prompt.startswith("/attach "):
            path = prompt[len("/attach "):].strip().strip('"')
            if len(attachments) >= ATTACH_MAX_FILES:
                print(f"[attach failed: max {ATTACH_MAX_FILES} files]")
                continue
            try:
                with open(path, "rb") as f:
                    data = f.read()
                import mimetypes
                mime = mimetypes.guess_type(path)[0] or ""
                attachments.extend(prepare_attachments([{
                    "name": os.path.basename(path),
                    "mime": mime,
                    "data": base64.b64encode(data).decode("ascii"),
                }]))
                names = ", ".join(a.name for a in attachments)
                print(f"[attached: {names} — sent with every turn this "
                      "session; /detach to clear]")
            except (OSError, ValueError) as e:
                print(f"[attach failed: {e}]")
            continue
        if prompt == "/detach":
            attachments = []
            print("[attachments cleared]")
            continue

        result = run_turn(
            clients, ctx, session_id, prompt,
            allow_code_execution=code_exec,
            attachments=attachments or None,
        )

        tier_summary = []
        if result.loaded_system:
            tier_summary.append(f"system:{len(result.loaded_system)}")
        if result.loaded_active:
            tier_summary.append(f"active:{','.join(result.loaded_active)}")
        if result.loaded_archive:
            tier_summary.append(f"archive:{','.join(result.loaded_archive)}")
        print(f"[loaded {' | '.join(tier_summary) if tier_summary else 'nothing'}]")

        print(f"\nagent> {result.answer}\n")

        if result.image and result.image.get("b64"):
            ext = (result.image.get("mime", "image/png").split("/")[-1]
                   or "png").split("+")[0]
            fname = f"generated-{datetime.now().strftime('%Y%m%d-%H%M%S')}.{ext}"
            with open(fname, "wb") as f:
                f.write(base64.b64decode(result.image["b64"]))
            print(f"[image saved: {os.path.abspath(fname)}]")
            if result.image.get("url"):
                print(f"[image url: {result.image['url']}]")
            print()

        if result.edits_applied:
            print(f"[memory: {'; '.join(result.edits_applied)}]\n")

    summarized = summarize_session_to_skill(clients, ctx, session_id)
    if summarized:
        print(f"[session saved: {summarized}]")


def _consolidate_one(user_id: str, dry_run: bool) -> None:
    ctx = UserCtx(user_id=user_id)
    print(f"\n=== Consolidating skills for {user_id} (dry_run={dry_run}) ===")
    result = consolidate(ctx, dry_run=dry_run)
    print(json.dumps(result, indent=2))


def _run_consolidate_cli() -> None:
    dry = "--dry-run" in sys.argv
    all_users = "--all-users" in sys.argv

    if all_users:
        uids = UserStore.list_user_ids()
        if not uids:
            print("No users found in the users table.", file=sys.stderr)
            sys.exit(1)
        for uid in uids:
            _consolidate_one(uid, dry)
        return

    user_id = os.environ.get("AGENT_USER_ID") or "default"
    try:
        validate_user_id(user_id)
    except ValueError as e:
        print(f"Invalid AGENT_USER_ID: {e}", file=sys.stderr)
        sys.exit(1)
    _consolidate_one(user_id, dry)


if __name__ == "__main__":
    main()
